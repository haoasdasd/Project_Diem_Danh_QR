from fastapi import FastAPI, Request, WebSocket
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from openpyxl import Workbook, load_workbook
from fastapi.staticfiles import StaticFiles
from datetime import datetime
from threading import Lock
import webbrowser
from pathlib import Path
import os
import json
import cv2
import numpy as np
import threading
import time
import socket
from fastapi.responses import HTMLResponse
from pathlib import Path


app = FastAPI()
lock = Lock()
app.mount("/static", StaticFiles(directory=".", html=True), name="static")



EXCEL_FILE = "Diem_Danh_Tong_Hop.xlsx"
QR_LOG_FILE = "qr_log.txt"
DISPLAY_LOG = "display_log.txt"
TOTAL_PEOPLE = 500

message_queue = []
display_thread_running = False
video_queue = []
video_thread_running = False
skip_all_videos = threading.Event()

pending_clients = set()
allowed_clients = set()
denied_clients = set()

connected_websockets = []

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    connected_websockets.append(websocket)
    try:
        while True:
            await websocket.receive_text()
    except:
        connected_websockets.remove(websocket)

async def notify_clients(message: str):
    for ws in connected_websockets:
        try:
            await ws.send_text(message)
        except:
            pass

PASSWORD = "Admin@123"

def ask_password():
    for _ in range(3):
        entered = input(" Nhập mật khẩu: ").strip()
        if entered == PASSWORD:
            print("  Mật khẩu chính xác.")
            return True
        else:
            print("  Sai mật khẩu.")
    print("  Nhập sai quá 3 lần. Thoát.")
    return False

def draw_info_box(frame, text, pos=None, box_size=(700, 40), color=(0, 255, 0)):
    h, w, _ = frame.shape
    box_w, box_h = box_size
    if pos is None:
        pos = (w - box_w - 10, h - 10)
    x, y = pos
    cv2.rectangle(frame, (x, y - box_h), (x + box_w, y), (0, 0, 0), -1)
    cv2.rectangle(frame, (x, y - box_h), (x + box_w, y), color, 2)
    cv2.putText(frame, text, (x + 10, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

def message_display_loop():
    global message_queue, display_thread_running
    display_thread_running = True
    while message_queue:
        message = message_queue.pop(0)
        h, w = 200, 800
        bg = np.zeros((h, w, 3), dtype=np.uint8)
        draw_info_box(bg, message, pos=(10, h - 10), color=(0, 255, 0))
        cv2.imshow("Server Overlay", bg)
        if cv2.waitKey(3000) & 0xFF == ord('q'):
            break
        cv2.destroyWindow("Server Overlay")
    display_thread_running = False

def show_overlay_message(text):
    global display_thread_running
    message_queue.append(text)
    if not display_thread_running:
        threading.Thread(target=message_display_loop, daemon=True).start()

def video_playback_loop():
    global video_queue, video_thread_running
    video_thread_running = True
    while video_queue:
        if skip_all_videos.is_set():
            video_queue.clear()
            skip_all_videos.clear()
            break

        video_path = video_queue.pop(0)
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            print(f"[x] Không mở được video: {video_path}")
            continue

        cv2.namedWindow("Video", cv2.WINDOW_NORMAL)
        cv2.setWindowProperty("Video", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

        while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break
            cv2.imshow("Video", frame)
            key = cv2.waitKey(30) & 0xFF
            if key == ord('q'):
                break
            if key == ord('n'):
                skip_all_videos.set()
                break

        cap.release()
        cv2.destroyWindow("Video")

        if skip_all_videos.is_set():
            video_queue.clear()
            skip_all_videos.clear()
            break

    video_thread_running = False

def init_files():
    print(" Kiểm tra các tệp hệ thống...")

    def confirm_reset(path, title, header):
        if os.path.exists(path):
            ans = input(f" Đã tồn tại '{path}'. Tạo mới? (y/n): ").strip().lower()
            if ans == "y":
                os.remove(path)
                if path.endswith(".xlsx"):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(header)
                    wb.save(path)
                else:
                    open(path, "w", encoding="utf-8").close()
                print(f" Tạo mới {title}.")
            else:
                print(f" Giữ lại {title} cũ.")
        else:
            if path.endswith(".xlsx"):
                wb = Workbook()
                ws = wb.active
                ws.append(header)
                wb.save(path)
            else:
                open(path, "w", encoding="utf-8").close()
            print(f" Tạo mới {title}.")

    confirm_reset(EXCEL_FILE, "file Excel", ["STT", "Tên", "Giới Tính", "Chức Vụ", "Đơn Vị", "Loại Đại Biểu", "Thời Gian"])
    confirm_reset(QR_LOG_FILE, "QR log", [])
    confirm_reset(DISPLAY_LOG, "display log", [])

def load_qr_log():
    with open(QR_LOG_FILE, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())

def save_qr_log(qr_data_dict):
    with open(QR_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(qr_data_dict, ensure_ascii=False) + "\n")

def write_display_log(stt, ten, qr_data_dict):
    message = f" STT {stt}: {ten} – Người thứ {stt}/{TOTAL_PEOPLE}"
    with open(DISPLAY_LOG, "a", encoding="utf-8") as logf:
        logf.write(json.dumps(qr_data_dict, ensure_ascii=False) + "\n")
    show_overlay_message(message)

@app.post("/request_access")
async def request_access(request: Request):
    client_ip = request.client.host
    if client_ip in allowed_clients:
        return {"status": "allowed"}
    elif client_ip in denied_clients:
        return {"status": "denied"}
    else:
        pending_clients.add(client_ip)
        await notify_clients(f"📡 Yêu cầu kết nối từ: {client_ip}")
        return {"status": "pending"}

@app.post("/scan")
async def scan(request: Request):
    ip = request.client.host
    if ip not in allowed_clients:
        return {"status": "denied", "message": " IP chưa được duyệt hoặc đã bị ngắt kết nối"}
    try:
        body = await request.json()
        ten = body.get("Ten", "")
        gioi_tinh = body.get("Gioi_Tinh", "")
        chuc_vu = body.get("Chuc_Vu", "")
        don_vi = body.get("Don_Vi", "")
        loai_dai_bieu = body.get("Loai_Dai_Bieu", "")
        ma_qr = body.get("Ma_QR", "").strip()
        video_path = body.get("video_path", "")

        if not ma_qr:
            return {"status": "error", "message": "Thiếu mã QR"}

        qr_data_dict = {
            "Tên": ten,
            "Giới Tính": gioi_tinh,
            "Chức Vụ": chuc_vu,
            "Đơn Vị": don_vi,
            "Loại Đại Biểu": loai_dai_bieu,
            "video_path": video_path
        }

        qr_data_json = json.dumps(qr_data_dict, ensure_ascii=False)

        with lock:
            da_quet = load_qr_log()
            if qr_data_json in da_quet:
                return {"status": "duplicate", "message": "Mã đã điểm danh trước đó"}

            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            stt = ws.max_row

            ws.append([
                stt,
                ten,
                gioi_tinh,
                chuc_vu,
                don_vi,
                loai_dai_bieu,
                now_str
            ])
            wb.save(EXCEL_FILE)

            save_qr_log(qr_data_dict)
            write_display_log(stt, ten, qr_data_dict)

            if video_path:
                video_queue.append(video_path)
                if not video_thread_running:
                    threading.Thread(target=video_playback_loop, daemon=True).start()

            return {
                "status": "success",
                "message": "Điểm danh thành công",
                "stt": stt,
                "video_path": video_path
            }

    except Exception as e:
        return {"status": "error", "message": f"Lỗi xử lý: {str(e)}"}

@app.get("/get_excel")
def get_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = [list(row) for row in rows]
    return JSONResponse(result)

@app.post("/add_excel")
async def add_excel(request: Request):
    data = await request.json()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    stt = ws.max_row
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        stt,
        data.get("ten", ""),
        data.get("gioi_tinh", ""),
        data.get("chuc_vu", ""),
        data.get("don_vi", ""),
        data.get("loai_db", ""),
        now_str
    ])
    wb.save(EXCEL_FILE)
    return {"status": "ok"}

@app.post("/update_excel")
async def update_excel(request: Request):
    body = await request.json()
    stt = int(body["stt"])
    col = int(body["col"])
    value = body["value"]
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == stt:
            row[col].value = value
            break
    wb.save(EXCEL_FILE)
    return {"status": "ok"}

@app.get("/get_qr_log")
def get_qr_log():
    try:
        if not os.path.exists("qr_log.txt"):
            return {"status": "ok", "lines": ["(Chưa có dữ liệu trong qr_log.txt)"]}
        with open("qr_log.txt", "r", encoding="utf-8") as f:
            lines = f.readlines()
        return {"status": "ok", "lines": lines}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/get_display_log")
def get_display_log():
    try:
        if not os.path.exists("display_log.txt"):
            return {"status": "ok", "lines": ["(Chưa có dữ liệu trong display_log.txt)"]}
        with open("display_log.txt", "r", encoding="utf-8") as f:
            lines = f.readlines()
        return {"status": "ok", "lines": lines}
    except Exception as e:
        return {"status": "error", "message": str(e)}



@app.post("/delete_excel")
async def delete_excel(request: Request):
    body = await request.json()
    stt = int(body["stt"])
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == stt:
            ws.delete_rows(i, 1)
            break
    wb.save(EXCEL_FILE)
    return {"status": "ok"}

@app.get("/get_clients")
def get_clients():
    return {"pending": list(pending_clients), "allowed": list(allowed_clients)}

@app.post("/approve_client")
async def approve_client(request: Request):
    data = await request.json()
    ip = data.get("ip")
    pending_clients.discard(ip)
    allowed_clients.add(ip)
    await notify_clients(f" Máy chủ đã duyệt {ip}")
    return {"status": "approved"}

@app.post("/deny_client")
async def deny_client(request: Request):
    data = await request.json()
    ip = data.get("ip")
    pending_clients.discard(ip)
    denied_clients.add(ip)
    await notify_clients(f" Máy chủ đã từ chối {ip}")
    return {"status": "denied"}

@app.post("/disconnect_client")
async def disconnect_client(request: Request):
    data = await request.json()
    ip = data.get("ip")
    allowed_clients.discard(ip)
    await notify_clients(f" Máy chủ đã ngắt kết nối {ip}")
    return {"status": "disconnected"}

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "localhost"

ip_address = get_local_ip()
print(f"📡 Máy chủ đang hoạt động tại http://{ip_address}:8000")

def open_dashboard():
    file_path = Path(__file__).parent / "dashboard.html"
    if file_path.exists():
        webbrowser.open(f"http://localhost:8000/dashboard.html")
    else:
        print("⚠️ Không tìm thấy file dashboard.html.")

@app.get("/dashboard.html", response_class=HTMLResponse)
def serve_dashboard():
    file_path = Path(__file__).parent / "dashboard.html"
    return file_path.read_text(encoding="utf-8")


def broadcast_server_ip():
    udp_socket = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    udp_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    udp_socket.bind(("", 9999))
    print("📡 Đang chờ client tìm kiếm máy chủ...")
    while True:
        try:
            data, addr = udp_socket.recvfrom(1024)
            if data.decode() == "DISCOVER_SERVER":
                ip = addr[0]
                print(f" Client yêu cầu kết nối: {ip}")
                pending_clients.add(ip)
                udp_socket.sendto(b"SERVER_IP_RESPONSE", addr)
        except Exception as e:
            print(f"[!] Lỗi broadcast: {e}")

threading.Thread(target=broadcast_server_ip, daemon=True).start()
print("Nhấn 'q' để bỏ qua 1 video _ Nhấn 'n' để bỏ qua toàn bộ video!!")

if __name__ == "__main__":
    import uvicorn
    import threading
    if not ask_password():
        exit()
    init_files()

    def start_server():
        print(f"[DEBUG] Danh sách file trong thư mục: {os.listdir(Path(__file__).parent)}")
        uvicorn.run(app, host="0.0.0.0", port=8000)

    threading.Thread(target=start_server, daemon=True).start()

    print(" Server đã khởi động.")
    time.sleep(2)
    open_dashboard()

    while True:
        print("\n MENU:")
        print("1. Mở lại giao diện quản lý")
        print("2. Thoát")
        choice = input("Chọn: ").strip()
        if choice == "1":
            open_dashboard()
        elif choice == "2":
            print(" Thoát ứng dụng.")
            break


