import cv2
import os
import threading
import time
from pyzbar.pyzbar import decode
from openpyxl import Workbook
from datetime import datetime

timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
excel_file = f"Diem_Danh_{timestamp}.xlsx"
try:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Tên", "Giới Tính", "Chức Vụ", "Đơn Vị", "Loại Đại Biểu", "Thời Gian Check in"])
    workbook.save(excel_file)
    print(f"Tạo file Excel mới: {excel_file}")
except Exception as e:
    print(f"Lỗi khi tạo file Excel: {e}")

current_video_thread = None
stop_video_flag = threading.Event()
scanned_qrs = set()
last_scan_time = 0
scan_delay = 3

def save_to_excel(data):
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        from openpyxl import load_workbook
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([
            data["Tên"], 
            data["Giới Tính"], 
            data["Chức Vụ"], 
            data["Đơn Vị"], 
            data["Loại Đại Biểu"], 
            now
        ])
        workbook.save(excel_file)
        print("Dữ liệu đã được lưu.")
    except Exception as e:
        print(f"Lỗi khi ghi vào Excel: {e}")

def play_video(video_path):
    global stop_video_flag
    try:
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            print(f"Lỗi khi mở video: {video_path}")
            return

        cv2.namedWindow("Video", cv2.WINDOW_NORMAL)
        cv2.setWindowProperty("Video", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

        while True:
            if stop_video_flag.is_set():
                break
            ret, frame = cap.read()
            if not ret:
                # Nếu đến cuối video, quay lại đầu video
                cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                continue

            cv2.imshow("Video", frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

        cap.release()
        cv2.destroyWindow("Video")
    except Exception as e:
        print(f"Lỗi khi phát video: {e}")

def process_qr_data(qr_data):
    global current_video_thread, stop_video_flag, scanned_qrs

    try:
        qr_info = eval(qr_data)
        video_path = qr_info["video_path"]

        if qr_data not in scanned_qrs:
            save_to_excel(qr_info)
            scanned_qrs.add(qr_data)
            print("QR mới được quét, thông tin đã lưu vào Excel.")
        else:
            print("QR đã được quét trước đó, không lưu vào Excel.")

        if current_video_thread and current_video_thread.is_alive():
            stop_video_flag.set()
            current_video_thread.join()

        stop_video_flag.clear()
        current_video_thread = threading.Thread(target=play_video, args=(video_path,))
        current_video_thread.start()

        print("Nội dung QR:", qr_info)

    except Exception as e:
        print(f"Lỗi khi xử lý dữ liệu QR: {e}")

def scan_qr_code():
    global last_scan_time

    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        for barcode in decode(frame):
            current_time = time.time()
            qr_data = barcode.data.decode("utf-8")

            if current_time - last_scan_time >= scan_delay:
                process_qr_data(qr_data)
                last_scan_time = current_time

        cv2.imshow("Quet Ma QR", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    print("Nhấn 'q' để thoát.")
    scan_qr_code()
